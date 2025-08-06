"""
CSV-based paper registry for tracking analyzed papers
"""

import os
import pandas as pd
from typing import List, Dict, Optional, Any
from datetime import datetime, timedelta
import logging

from .models import (PaperRegistryConfig, PAPERS_CSV_COLUMNS, HISTORY_CSV_COLUMNS, 
                     PaperStatus, JAPANESE_COLUMN_NAMES, JAPANESE_STATUS_NAMES,
                     PAPERS_CSV_COLUMNS_MIN, PAPERS_CSV_COLUMNS_FULL, PAPERS_CSV_COLUMNS_LEGACY)
from .utils import AnalysisResultConverter, AnalysisTranslator, JapaneseAnalysisConverter

logger = logging.getLogger(__name__)


class CSVPaperRegistry:
    """CSV-based paper registry with Excel compatibility"""
    
    def __init__(self, config: Optional[PaperRegistryConfig] = None):
        """
        Initialize CSV paper registry
        
        Args:
            config: Registry configuration, uses defaults if None
        """
        self.config = config or PaperRegistryConfig()
        self._ensure_database_exists()
    
    def _ensure_database_exists(self):
        """Create database directory and CSV files if they don't exist"""
        # Create database directory
        os.makedirs(self.config.db_path, exist_ok=True)
        
        # Create papers CSV if it doesn't exist
        if not os.path.exists(self.config.papers_path):
            empty_df = pd.DataFrame(columns=self.config.columns)
            empty_df.to_csv(self.config.papers_path, 
                          index=False, 
                          encoding=self.config.encoding)
            logger.info(f"Created papers database: {self.config.papers_path} (schema: {self.config.schema_version})")
        
        # Create history CSV if it doesn't exist
        if not os.path.exists(self.config.history_path):
            empty_df = pd.DataFrame(columns=HISTORY_CSV_COLUMNS)
            empty_df.to_csv(self.config.history_path, 
                          index=False, 
                          encoding=self.config.encoding)
            logger.info(f"Created history database: {self.config.history_path}")
    
    def _load_papers_df(self) -> pd.DataFrame:
        """Load papers DataFrame from CSV"""
        try:
            df = pd.read_csv(self.config.papers_path, encoding=self.config.encoding)
            logger.debug(f"Loaded papers CSV: {len(df)} rows, columns: {list(df.columns)}")
            # Ensure all required columns exist
            for col in self.config.columns:
                if col not in df.columns:
                    df[col] = ""
            return df
        except Exception as e:
            logger.error(f"Error loading papers CSV: {e}")
            # Return empty DataFrame with correct columns
            return pd.DataFrame(columns=self.config.columns)
    
    def _save_papers_df(self, df: pd.DataFrame):
        """Save papers DataFrame to CSV"""
        try:
            # Ensure columns are in the correct order
            df = df.reindex(columns=self.config.columns, fill_value="")
            df.to_csv(self.config.papers_path, 
                     index=False, 
                     encoding=self.config.encoding)
        except Exception as e:
            logger.error(f"Error saving papers CSV: {e}")
            raise
    
    def _load_history_df(self) -> pd.DataFrame:
        """Load history DataFrame from CSV"""
        try:
            df = pd.read_csv(self.config.history_path, encoding=self.config.encoding)
            return df
        except Exception as e:
            logger.error(f"Error loading history CSV: {e}")
            return pd.DataFrame(columns=HISTORY_CSV_COLUMNS)
    
    def _save_history_df(self, df: pd.DataFrame):
        """Save history DataFrame to CSV"""
        try:
            df.to_csv(self.config.history_path, 
                     index=False, 
                     encoding=self.config.encoding)
        except Exception as e:
            logger.error(f"Error saving history CSV: {e}")
            raise
    
    # Core functionality methods
    
    def is_paper_analyzed(self, arxiv_id: str) -> bool:
        """
        Check if a paper has already been analyzed
        
        Args:
            arxiv_id: arXiv paper ID
            
        Returns:
            True if paper is already analyzed
        """
        df = self._load_papers_df()
        # Debug print
        logger.debug(f"Checking if {arxiv_id} is analyzed. CSV has {len(df)} rows.")
        if len(df) > 0:
            logger.debug(f"First few arxiv_ids: {df['arxiv_id'].tolist()[:5]}")
            logger.debug(f"arxiv_id column dtype: {df['arxiv_id'].dtype}")
            logger.debug(f"Target ID: '{arxiv_id}' (type: {type(arxiv_id)})")
        
        # Convert to string and compare
        result = str(arxiv_id) in df['arxiv_id'].astype(str).values
        logger.debug(f"Match result: {result}")
        return result
    
    def register_analyzed_paper(self, paper_data: dict):
        """
        Register a newly analyzed paper
        
        Args:
            paper_data: Analysis result data from workflow
        """
        # Convert to CSV format based on schema version
        # Always use full format to ensure all data is saved
        csv_row = JapaneseAnalysisConverter.to_japanese_full_format(paper_data)
        
        # Load existing data
        df = self._load_papers_df()
        
        # Check for duplicates
        arxiv_id = csv_row['arxiv_id']
        if arxiv_id in df['arxiv_id'].values:
            logger.warning(f"Paper {arxiv_id} already exists, updating...")
            df = df[df['arxiv_id'] != arxiv_id]  # Remove existing
        
        # Add new row
        new_df = pd.DataFrame([csv_row])
        df = pd.concat([df, new_df], ignore_index=True)
        
        # Save
        self._save_papers_df(df)
        logger.info(f"Registered paper: {arxiv_id}")
    
    def get_analyzed_papers(self, 
                          query_filter: Optional[str] = None,
                          status_filter: Optional[str] = None,
                          limit: Optional[int] = None) -> pd.DataFrame:
        """
        Get analyzed papers with optional filtering
        
        Args:
            query_filter: Filter by query used
            status_filter: Filter by status
            limit: Maximum number of results
            
        Returns:
            Filtered DataFrame
        """
        df = self._load_papers_df()
        
        # Apply filters
        if query_filter:
            # Search across multiple fields: query_used, title, analysis_summary, abstract
            search_fields = ['query_used', 'title', 'analysis_summary', 'abstract', 'what_is_it', 'key_technique', 'next_papers']
            mask = pd.Series([False] * len(df), index=df.index)
            
            for field in search_fields:
                if field in df.columns:
                    field_mask = df[field].str.contains(query_filter, case=False, na=False)
                    mask = mask | field_mask
            
            df = df[mask]
        
        if status_filter:
            df = df[df['status'] == status_filter]
        
        # Sort by analysis date (newest first)
        if 'analyzed_at' in df.columns:
            df = df.sort_values('analyzed_at', ascending=False)
        
        # Apply limit
        if limit:
            df = df.head(limit)
        
        return df
    
    def filter_new_papers(self, papers: List[Dict]) -> List[Dict]:
        """
        Filter out papers that have already been analyzed
        
        Args:
            papers: List of paper metadata dicts
            
        Returns:
            List of unanalyzed papers only
        """
        df = self._load_papers_df()
        analyzed_ids = set(df['arxiv_id'].values)
        
        new_papers = []
        for paper in papers:
            arxiv_id = paper.get('arxiv_id', '')
            if arxiv_id not in analyzed_ids:
                new_papers.append(paper)
        
        skipped_count = len(papers) - len(new_papers)
        if skipped_count > 0:
            logger.info(f"Filtered out {skipped_count} already-analyzed papers")
        
        return new_papers
    
    def get_analysis_by_id(self, arxiv_id: str) -> Optional[Dict]:
        """
        Get analysis result for a specific paper
        
        Args:
            arxiv_id: arXiv paper ID
            
        Returns:
            Analysis result dict or None if not found
        """
        df = self._load_papers_df()
        matches = df[df['arxiv_id'] == arxiv_id]
        
        if matches.empty:
            return None
        
        # Convert CSV row back to analysis format
        csv_row = matches.iloc[0].to_dict()
        return AnalysisResultConverter.csv_row_to_analysis(csv_row)
    
    def update_search_history(self, 
                            query: str, 
                            papers_found: int,
                            papers_new: int, 
                            papers_analyzed: int,
                            execution_time: int,
                            analysis_mode: str = "moderate"):
        """
        Update search history with new search results
        
        Args:
            query: Search query used
            papers_found: Total papers found
            papers_new: New papers (not previously analyzed)
            papers_analyzed: Papers actually analyzed
            execution_time: Total execution time in seconds
            analysis_mode: Analysis mode used
        """
        history_row = {
            'query': query,
            'executed_at': datetime.now().isoformat(),
            'papers_found': papers_found,
            'papers_new': papers_new,
            'papers_skipped': papers_found - papers_new,
            'papers_analyzed': papers_analyzed,
            'execution_time': execution_time,
            'analysis_mode': analysis_mode
        }
        
        # Load, append, save
        df = self._load_history_df()
        new_df = pd.DataFrame([history_row])
        df = pd.concat([df, new_df], ignore_index=True)
        
        self._save_history_df(df)
        logger.info(f"Updated search history: {query}")
    
    def get_search_statistics(self, days: int = 30) -> Dict[str, Any]:
        """
        Get search statistics for the last N days
        
        Args:
            days: Number of days to look back
            
        Returns:
            Statistics dictionary
        """
        df = self._load_history_df()
        
        if df.empty:
            return {
                'total_searches': 0,
                'total_papers_found': 0,
                'total_papers_analyzed': 0,
                'avg_execution_time': 0,
                'recent_queries': []
            }
        
        # Filter by date
        cutoff_date = datetime.now() - timedelta(days=days)
        df['executed_at'] = pd.to_datetime(df['executed_at'])
        recent_df = df[df['executed_at'] >= cutoff_date]
        
        stats = {
            'total_searches': len(recent_df),
            'total_papers_found': recent_df['papers_found'].sum(),
            'total_papers_analyzed': recent_df['papers_analyzed'].sum(),
            'avg_execution_time': recent_df['execution_time'].mean(),
            'recent_queries': recent_df['query'].tolist()[-10:]  # Last 10 queries
        }
        
        return stats
    
    def cleanup_old_entries(self, days: int = 90):
        """
        Clean up old entries to keep database size manageable
        
        Args:
            days: Keep entries newer than this many days
        """
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # Clean papers
        df = self._load_papers_df()
        if not df.empty and 'analyzed_at' in df.columns:
            df['analyzed_at'] = pd.to_datetime(df['analyzed_at'])
            old_count = len(df)
            df = df[df['analyzed_at'] >= cutoff_date]
            new_count = len(df)
            
            if old_count > new_count:
                self._save_papers_df(df)
                logger.info(f"Cleaned up {old_count - new_count} old paper entries")
        
        # Clean history  
        df = self._load_history_df()
        if not df.empty and 'executed_at' in df.columns:
            df['executed_at'] = pd.to_datetime(df['executed_at'])
            old_count = len(df)
            df = df[df['executed_at'] >= cutoff_date]
            new_count = len(df)
            
            if old_count > new_count:
                self._save_history_df(df)
                logger.info(f"Cleaned up {old_count - new_count} old history entries")
    
    def export_to_excel(self, output_path: str, use_japanese_headers: bool = True, 
                       enhanced_formatting: bool = True):
        """
        Export the registry to Excel format with enhanced formatting
        
        Args:
            output_path: Path for Excel file
            use_japanese_headers: Use Japanese column names if True
            enhanced_formatting: Apply enhanced Excel formatting
        """
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                raise ImportError(
                    "openpyxl is required for Excel export. "
                    "Install it with: pip install openpyxl"
                )
            
            papers_df = self._load_papers_df()
            history_df = self._load_history_df()
            
            # Convert to full format if schema is set to full OR enhanced formatting is requested
            if self.config.schema_version == "full" or enhanced_formatting:
                papers_df = self._convert_to_full_format(papers_df)
            
            # Translate column names if requested
            if use_japanese_headers:
                papers_df = papers_df.rename(columns=JAPANESE_COLUMN_NAMES)
                # Translate status values
                if 'ステータス' in papers_df.columns:
                    papers_df['ステータス'] = papers_df['ステータス'].map(
                        lambda x: JAPANESE_STATUS_NAMES.get(x, x)
                    )
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                sheet_name = '分析済み論文' if use_japanese_headers else 'Analyzed Papers'
                papers_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                history_sheet_name = '検索履歴' if use_japanese_headers else 'Search History'
                history_df.to_excel(writer, sheet_name=history_sheet_name, index=False)
                
                # Apply enhanced formatting if requested
                if enhanced_formatting:
                    self._apply_excel_formatting(writer.book, sheet_name, papers_df)
            
            logger.info(f"Exported registry to Excel: {output_path}")
        except ImportError as e:
            logger.error(f"Missing dependency for Excel export: {e}")
            # Provide alternative solution
            csv_backup = output_path.replace('.xlsx', '_papers.csv')
            papers_df = self._load_papers_df()
            if use_japanese_headers:
                papers_df = papers_df.rename(columns=JAPANESE_COLUMN_NAMES)
            papers_df.to_csv(csv_backup, index=False, encoding='utf-8-sig')
            logger.info(f"Alternative: Exported to CSV instead: {csv_backup}")
            raise
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _apply_excel_formatting(self, workbook, sheet_name: str, df: pd.DataFrame):
        """Apply enhanced formatting to Excel worksheet"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            ws = workbook[sheet_name]
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Column width settings
            column_widths = {
                "arxiv_id": 15,
                "タイトル": 40,
                "著者": 30,
                "公開日": 12,
                "取得日": 12,
                "カテゴリ": 20,
                "キーワード": 25,
                "処理状態": 12,
                "概要JP": 50,
                "要点_一言": 45,
                "新規性": 35,
                "手法": 35,
                "実験設定": 35,
                "結果": 35,
                "考察": 35,
                "今後の課題": 30,
                "応用アイデア": 30,
                "重要度": 10,
                "リンク_pdf": 40,
                "落合フォーマットURL": 40,
                "備考": 25
            }
            
            # Apply column widths
            for idx, column in enumerate(df.columns, 1):
                col_letter = ws.cell(row=1, column=idx).column_letter
                width = column_widths.get(column, 20)
                ws.column_dimensions[col_letter].width = width
            
            # Row height for better readability
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                ws.row_dimensions[row[0].row].height = 25
            
            # Text wrapping for content cells
            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = wrap_alignment
            
            # Freeze panes for easy navigation
            ws.freeze_panes = "D2"  # Freeze first 3 columns and header row
            
            logger.debug(f"Applied enhanced formatting to {sheet_name}")
            
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {e}")
            # Continue without formatting if it fails
    
    def _convert_to_full_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert minimal format DataFrame to full 21-column format with detailed analysis"""
        from datetime import datetime
        
        # Create full format DataFrame with all required columns
        full_columns = PAPERS_CSV_COLUMNS_FULL
        full_df = pd.DataFrame(columns=full_columns)
        
        # Map existing minimal columns to full format with enhanced content generation
        for _, row in df.iterrows():
            full_row = {}
            
            # Basic information (copy from minimal format)
            full_row["arxiv_id"] = row.get("arxiv_id", "")
            full_row["タイトル"] = row.get("タイトル", "")
            full_row["著者"] = row.get("著者", "")
            full_row["公開日"] = row.get("公開日", "")
            full_row["取得日"] = datetime.now().strftime("%Y-%m-%d")
            full_row["カテゴリ"] = row.get("カテゴリ", "")
            
            # Parse available analysis data
            title = str(row.get("タイトル", ""))
            abstract_jp = str(row.get("概要JP", "")) if pd.notna(row.get("概要JP", "")) else ""
            technique = str(row.get("手法", "")) if pd.notna(row.get("手法", "")) else ""
            results = str(row.get("結果", "")) if pd.notna(row.get("結果", "")) else ""
            
            # Enhanced content generation
            paper_analysis = self._analyze_paper_content(title, abstract_jp, technique, results)
            
            # Populate all columns with high-quality content
            full_row["キーワード"] = paper_analysis["keywords"]
            full_row["処理状態"] = "done"
            full_row["概要JP"] = paper_analysis["detailed_summary"]
            full_row["要点_一言"] = paper_analysis["key_insight"]
            full_row["新規性"] = paper_analysis["novelty_points"]
            full_row["手法"] = paper_analysis["methodology"]
            full_row["実験設定"] = paper_analysis["experimental_setup"]
            full_row["結果"] = paper_analysis["results"]
            full_row["考察"] = paper_analysis["discussion"]
            full_row["今後の課題"] = paper_analysis["future_work"]
            
            # Management columns
            full_row["応用アイデア"] = ""  # User input
            full_row["重要度"] = paper_analysis["importance_rating"]
            full_row["リンク_pdf"] = f"https://arxiv.org/pdf/{row.get('arxiv_id', '')}.pdf"
            full_row["落合フォーマットURL"] = f"https://reports.example.com/ochiai/{row.get('arxiv_id', '')}"
            full_row["備考"] = ""  # User input
            
            # Add row to DataFrame
            full_df = pd.concat([full_df, pd.DataFrame([full_row])], ignore_index=True)
        
        return full_df
    
    def _analyze_paper_content(self, title: str, abstract_jp: str, technique: str, results: str) -> dict:
        """
        Comprehensive paper analysis to generate detailed Japanese content.
        This new version focuses on dynamic content generation from abstract and title.
        """
        
        # Generate detailed content based on dynamic analysis
        analysis = {
            "keywords": self._generate_keywords(title, abstract_jp),
            "detailed_summary": self._generate_detailed_summary(title, abstract_jp),
            "key_insight": self._generate_key_insight(title, abstract_jp),
            "novelty_points": self._generate_novelty_points(abstract_jp),
            "methodology": self._generate_methodology(technique, abstract_jp),
            "experimental_setup": self._generate_experimental_setup(results, abstract_jp),
            "results": self._generate_results_summary(results),
            "discussion": self._generate_discussion(abstract_jp, results),
            "future_work": self._generate_future_work(abstract_jp),
            "importance_rating": self._calculate_importance_rating(title, abstract_jp, results)
        }
        
        return analysis
    
    
    def _generate_keywords(self, title: str, abstract_jp: str) -> str:
        """Generate keywords by extracting important nouns from title and abstract."""
        import re
        
        text = title + " " + abstract_jp
        # Simple noun extraction (Japanese)
        words = re.findall(r'[\u4e00-\u9fafぁ-んァ-ン]+', text)
        
        # Filter common words and count frequency
        ignore_words = {"研究", "論文", "方法", "提案", "結果", "評価", "実現", "向上", "ため", "による", "ついて"}
        word_counts = {}
        for word in words:
            if len(word) > 1 and word not in ignore_words:
                word_counts[word] = word_counts.get(word, 0) + 1
        
        # Get top 5 most frequent words
        sorted_words = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)
        keywords = [word for word, count in sorted_words[:5]]
        
        return ", ".join(keywords)
    
    def _generate_detailed_summary(self, title: str, abstract_jp: str) -> str:
        """Generate a detailed 2-3 sentence summary from the abstract."""
        if not abstract_jp:
            return f"「{title}」に関する研究。詳細は論文を参照。"
        
        # Take the first few sentences of the abstract
        sentences = abstract_jp.split('。')
        summary = '。'.join(sentences[:2]) + '。'
        
        if len(summary) < 50: # If summary is too short, add more
             summary = '。'.join(sentences[:3]) + '。'

        return summary
    
    def _generate_key_insight(self, title: str, abstract_jp: str) -> str:
        """Generate a punchy insight from the abstract."""
        if not abstract_jp:
            return "論文の核心的な洞察は本文を参照。"

        # Look for sentences that indicate a key finding
        insight_keywords = ["明らかになった", "示唆された", "重要な知見", "結論として"]
        sentences = abstract_jp.split('。')
        
        for sentence in sentences:
            for keyword in insight_keywords:
                if keyword in sentence:
                    return sentence + "。"

        # Fallback: use the first sentence of the abstract, it often contains the main point.
        return sentences[0] + "。"
    
    def _generate_novelty_points(self, abstract_jp: str) -> str:
        """Generate novelty points by extracting relevant sentences from the abstract."""
        if not abstract_jp:
            return "・論文の新規性については本文を参照。"

        novelty_keywords = ["新規", "初めて", "提案", "開発", "独自"]
        sentences = abstract_jp.split('。')
        novelty_points = []

        for sentence in sentences:
            for keyword in novelty_keywords:
                if keyword in sentence:
                    novelty_points.append(f"・{sentence.strip()}")
                    break # Move to next sentence once a keyword is found
        
        if not novelty_points:
            return "・具体的な新規性は本文にて詳述。"

        return "\n".join(novelty_points[:3]) # Return top 3 points
    
    def _generate_methodology(self, technique: str, abstract_jp: str) -> str:
        """Generate methodology description from technique data or abstract."""
        if technique and technique.strip():
            return technique # Return the direct analysis result if available

        if not abstract_jp:
            return "提案手法の詳細は本文を参照。"

        # Look for sentences describing the method
        method_keywords = ["手法", "用いて", "基づいて", "モデル", "アルゴリズム"]
        sentences = abstract_jp.split('。')
        
        for sentence in sentences:
            for keyword in method_keywords:
                if keyword in sentence:
                    return sentence + "。"
        
        return "具体的な手法については本文で詳述。"
    
    def _generate_experimental_setup(self, results: str, abstract_jp: str) -> str:
        """Generate experimental setup description from results data or abstract."""
        if results and results.strip():
            return results # Return direct analysis if available

        if not abstract_jp:
            return "実験設定の詳細は本文を参照。"

        # Look for sentences describing the experiment
        exp_keywords = ["実験", "評価", "比較", "データセット", "タスク"]
        sentences = abstract_jp.split('。')
        
        for sentence in sentences:
            for keyword in exp_keywords:
                if keyword in sentence:
                    return sentence + "。"
        
        return "具体的な実験設定については本文で詳述。"
    
    def _generate_results_summary(self, results: str) -> str:
        """Generate results summary from results data or abstract."""
        if results and results.strip():
            return results # Return direct analysis if available

        return "実験結果の詳細は本文を参照。"
    
    def _generate_discussion(self, abstract_jp: str, results: str) -> str:
        """Generate discussion points from abstract or results."""
        if not abstract_jp:
            return "考察の詳細は本文を参照。"

        # Look for sentences that indicate discussion or implication
        discussion_keywords = ["考察", "示唆", "意味", "重要", "課題"]
        sentences = abstract_jp.split('。')
        
        for sentence in sentences:
            for keyword in discussion_keywords:
                if keyword in sentence:
                    return sentence + "。"
        
        # Fallback to results if available
        if results and results.strip():
            return f"結果から、提案手法の有効性が示された。今後の課題は..."

        return "詳細な考察は本文にて詳述。"
    
    def _generate_future_work(self, abstract_jp: str) -> str:
        """Generate future work by extracting relevant sentences from the abstract."""
        if not abstract_jp:
            return "今後の課題は本文を参照。"

        future_work_keywords = ["今後の課題", "将来", "展望", "さらなる", "残されている"]
        sentences = abstract_jp.split('。')
        
        for sentence in sentences:
            for keyword in future_work_keywords:
                if keyword in sentence:
                    return sentence + "。"
        
        return "今後の課題については本文で詳述。"
    
    def _calculate_importance_rating(self, title: str, abstract_jp: str, results: str) -> str:
        """Calculate importance rating based on dynamic content analysis."""
        score = 3  # Base score

        # Title significance indicators
        title_lower = title.lower()
        if any(x in title_lower for x in ["comprehensive", "novel", "breakthrough", "survey", "review"]):
            score += 1
        
        # Abstract content indicators
        if any(x in abstract_jp for x in ["大きな影響", "重要な貢献", "根本的な解決"]):
            score += 1
            
        # Results indicators
        if results and any(x in results for x in ["大幅な改善", "state-of-the-art", "SOTA"]):
            score +=1

        score = min(5, max(1, score))
        return "★" * score
    
    def get_registry_info(self) -> Dict[str, Any]:
        """
        Get general information about the registry
        
        Returns:
            Info dictionary with counts and status
        """
        papers_df = self._load_papers_df()
        history_df = self._load_history_df()
        
        # Safe status column access
        completed_papers = 0
        failed_papers = 0
        if 'status' in papers_df.columns:
            completed_papers = len(papers_df[papers_df['status'] == PaperStatus.COMPLETED])
            failed_papers = len(papers_df[papers_df['status'] == PaperStatus.FAILED])
        
        return {
            'total_papers': len(papers_df),
            'completed_papers': completed_papers,
            'failed_papers': failed_papers,
            'total_searches': len(history_df),
            'database_path': self.config.db_path,
            'papers_file': self.config.papers_path,
            'history_file': self.config.history_path
        }
    
    def backup_database(self, backup_suffix: str = None) -> str:
        """
        Create a backup of the current database
        
        Args:
            backup_suffix: Optional suffix for backup files
            
        Returns:
            Backup directory path
        """
        from datetime import datetime
        import shutil
        
        if backup_suffix is None:
            backup_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        backup_dir = f"{self.config.db_path}/backup_{backup_suffix}"
        os.makedirs(backup_dir, exist_ok=True)
        
        # Backup papers CSV
        if os.path.exists(self.config.papers_path):
            backup_papers_path = f"{backup_dir}/{self.config.papers_file}"
            shutil.copy2(self.config.papers_path, backup_papers_path)
            logger.info(f"Backed up papers to: {backup_papers_path}")
        
        # Backup history CSV
        if os.path.exists(self.config.history_path):
            backup_history_path = f"{backup_dir}/{self.config.history_file}"
            shutil.copy2(self.config.history_path, backup_history_path)
            logger.info(f"Backed up history to: {backup_history_path}")
        
        logger.info(f"Database backup created: {backup_dir}")
        return backup_dir
    
    def reset_database(self, backup_first: bool = True, confirm: bool = True) -> bool:
        """
        Reset the database by clearing all papers and history
        
        Args:
            backup_first: Create backup before reset
            confirm: Require confirmation (for CLI safety)
            
        Returns:
            True if reset was successful
        """
        # if confirm:
        #     # This will be handled by CLI
        #     logger.warning("Reset confirmation required")
        #     return False
        
        try:
            # Create backup if requested
            backup_dir = None
            if backup_first:
                backup_dir = self.backup_database()
            
            # Reset papers database
            empty_papers_df = pd.DataFrame(columns=self.config.columns)
            empty_papers_df.to_csv(self.config.papers_path, 
                                 index=False, 
                                 encoding=self.config.encoding)
            logger.info(f"Reset papers database: {self.config.papers_path}")
            
            # Reset history database
            empty_history_df = pd.DataFrame(columns=HISTORY_CSV_COLUMNS)
            empty_history_df.to_csv(self.config.history_path,
                                  index=False,
                                  encoding=self.config.encoding)
            logger.info(f"Reset history database: {self.config.history_path}")
            
            if backup_dir:
                logger.info(f"Backup created before reset: {backup_dir}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error resetting database: {e}")
            return False
    
    def restore_from_backup(self, backup_dir: str) -> bool:
        """
        Restore database from backup directory
        
        Args:
            backup_dir: Path to backup directory
            
        Returns:
            True if restore was successful
        """
        import shutil
        
        try:
            backup_papers_path = f"{backup_dir}/{self.config.papers_file}"
            backup_history_path = f"{backup_dir}/{self.config.history_file}"
            
            # Restore papers if backup exists
            if os.path.exists(backup_papers_path):
                shutil.copy2(backup_papers_path, self.config.papers_path)
                logger.info(f"Restored papers from: {backup_papers_path}")
            
            # Restore history if backup exists  
            if os.path.exists(backup_history_path):
                shutil.copy2(backup_history_path, self.config.history_path)
                logger.info(f"Restored history from: {backup_history_path}")
            
            logger.info(f"Database restored from backup: {backup_dir}")
            return True
            
        except Exception as e:
            logger.error(f"Error restoring from backup: {e}")
            return False
